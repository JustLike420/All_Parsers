import csv
import json

from openpyxl import load_workbook, Workbook


def main():
    with open('data/json_data_ja1.json') as json_file:
        data = json.load(json_file)

    d = {int(k): v for k, v in data.items()}
    sorted_data = dict(sorted(d.items()))
    stream = open(f"data/JA1.csv", "w", newline='', encoding='utf-8')
    writer = csv.writer(stream)
    flag = True
    for lvl, words in sorted_data.items():
        glob_lvl = 0
        for word in words:
            if lvl != glob_lvl and word != '':
                glob_lvl = lvl
                if flag:
                    flag = False
                else:
                    writer.writerow(['', ''])
                writer.writerow([glob_lvl, word])
            elif word != '':
                writer.writerow(['', word])
        print('write', lvl)

def xlsx_converter():
    wb = Workbook()
    wb = load_workbook('rezult1.xlsx')
    de_ws = wb.create_sheet()
    de_ws.title = 'rezult_DE'
    br_ws = wb.create_sheet()
    br_ws.title = 'rezult_BR'
    es_ws = wb.create_sheet()
    es_ws.title = 'rezult_ES'
    fr_ws = wb.create_sheet()
    fr_ws.title = 'rezult_FR'
    tr_ws = wb.create_sheet()
    tr_ws.title = 'rezult_TR'
    ja_ws = wb.create_sheet()
    ja_ws.title = 'rezult_JA'
    with open(f'data/DE1.csv', 'r', encoding="utf-8") as f:
        for row in csv.reader(f):
            de_ws.append(row)
    with open(f'data/BR1.csv', 'r', encoding="utf-8") as f:
        for row in csv.reader(f):
            br_ws.append(row)
    with open(f'data/ES1.csv', 'r', encoding="utf-8") as f:
        for row in csv.reader(f):
            es_ws.append(row)
    with open(f'data/FR1.csv', 'r', encoding="utf-8") as f:
        for row in csv.reader(f):
            fr_ws.append(row)
    with open(f'data/TR1.csv', 'r', encoding="utf-8") as f:
        for row in csv.reader(f):
            tr_ws.append(row)
    with open(f'data/JA1.csv', 'r', encoding="utf-8") as f:
        for row in csv.reader(f):
            ja_ws.append(row)

    wb.save(f'rezult1.xlsx')


if __name__ == '__main__':
    main()
    xlsx_converter()

