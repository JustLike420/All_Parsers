import csv
import json

from openpyxl import load_workbook, Workbook


def csv_convertor():
    # langs = ['en', 'de', 'br', 'es', 'fr', 'ja']
    langs = ['tr']
    for lang in langs:
        with open(f'{lang}_data.json') as json_file:
            data = json.load(json_file)
        print(len(data.keys()))
        stream = open(f"data/{lang}.csv", "w", newline='', encoding='utf-8')
        writer = csv.writer(stream)

        d = {int(k): v for k, v in data.items()}
        sorted_data = dict(sorted(d.items()))

        for lvl, info in sorted_data.items():

            writer.writerow([lvl, info['title'], info['words'][0]])
            for word in info['words'][1:]:
                writer.writerow(['', '', word])
            writer.writerow(['', '', ''])
        print(data['1'])


def xlsx_convert():
    wb = Workbook()
    langs = ['en', 'de', 'br', 'es', 'fr', 'ja', 'tr']
    for lang in langs:
        ws = wb.create_sheet()
        ws.title = f'rezult_{lang}'
        with open(f'data/{lang}.csv', 'r', encoding="utf-8") as f:
            for row in csv.reader(f):
                ws.append(row)
    wb.save("data.xlsx")

if __name__ == '__main__':
    xlsx_convert()
    # csv_convertor()
