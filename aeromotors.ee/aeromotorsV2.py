import os

import openpyxl
import requests
from bs4 import BeautifulSoup
import math

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/94.0.4606.61 Safari/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9'
}
new_book = openpyxl.Workbook()
new_sheet = new_book.active
new_sheet['A1'].value = 'Бренд'
new_sheet['B1'].value = 'Наименование'
new_sheet['C1'].value = 'Артикул'
new_sheet['D1'].value = 'Цена'
new_sheet['E1'].value = 'Срок'
new_sheet['F1'].value = 'Количество'


def search_code(code):
    url = f'https://aeromotors.ee/ru/kataloog/search?search_query={code}'
    urls = []
    req = requests.get(url, headers=headers)
    src = req.text
    soup = BeautifulSoup(src, 'lxml')
    try:
        total = int(soup.find('div', class_='totaltext').find('b').text)
    except:
        total = 1
    blocks = soup.find_all('div', class_='tdlist-row')
    if blocks != 0:
        for block in blocks:
            urls.append(block.find('a', class_='article-name').get('href'))

    if total > 20:
        pages = math.ceil(total / 20)
        for page_num in range(2, pages+1):
            url = f'https://aeromotors.ee/ru/kataloog/search?search_query={code}&page={page_num}'
            req = requests.get(url, headers=headers)
            src = req.text
            soup = BeautifulSoup(src, 'lxml')
            blocks = soup.find_all('div', class_='tdlist-row')
            if blocks != 0:
                for block in blocks:
                    urls.append(block.find('a', class_='article-name').get('href'))
    return urls


def parse_card(url):
    lines = []
    req = requests.get(url, headers=headers)
    src = req.text
    soup = BeautifulSoup(src, 'lxml')

    brand = soup.find('meta', itemprop='brand').get('content')
    card_code = soup.find('meta', itemprop='mpn').get('content')
    name = soup.find('h1', itemprop='name').text.strip().replace(brand + ' ', '').replace(' ' + card_code, '')

    containers = soup.find_all('div', class_='dailyprice-container')

    for container in containers:
        price = container.find('div', class_='price-container').text.strip()
        date = container.find('div', class_='delivery-container').text.strip()
        count = container.find('div', class_='quantity-container').text.strip()
        new_sheet.append((brand, name, card_code, price, date, count))
        new_book.save('data3.xlsx')


if __name__ == '__main__':
    table = openpyxl.open(f'data2.xlsx')
    sheet = table.active
    if os.path.exists('table_row.txt'):
        with open('table_row.txt', 'r+', encoding='utf-8') as log_file:
            start_row = int(log_file.readlines()[0])
    else:
        start_row = 2
    for row in range(start_row, sheet.max_row + 1):
        code = sheet[row][0].value
        if code is not None:
            urls = search_code(code)
            print(code, f'{len(urls)}', 'start!')
            for i, url in enumerate(urls):
                parse_card(url)
            with open('table_row.txt', 'w+', encoding='utf-8') as log_file:
                log_file.write(str(row))
            print(code, f'{len(urls)}', 'finish!')
        print(f'{row} / {sheet.max_row}')
